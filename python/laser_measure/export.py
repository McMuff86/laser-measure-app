"""
Export functionality for laser measurements.

Supports exporting measurement data to CSV, JSON, and Excel formats.
Includes special formatting for door measurements workflow.
"""

import csv
import json
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
from dataclasses import asdict

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from .protocol import MeasurementResult


class MeasurementExporter:
    """Export measurements to various formats"""
    
    @staticmethod
    def to_csv(measurements: List[MeasurementResult], 
               filepath: str,
               door_format: bool = False) -> bool:
        """Export measurements to CSV format
        
        Args:
            measurements: List of measurement results
            filepath: Output CSV file path
            door_format: Use door-specific formatting (3 measurements per door)
            
        Returns:
            True if export successful
        """
        try:
            path = Path(filepath)
            path.parent.mkdir(parents=True, exist_ok=True)
            
            with open(path, 'w', newline='', encoding='utf-8') as csvfile:
                if door_format:
                    MeasurementExporter._write_door_csv(csvfile, measurements)
                else:
                    MeasurementExporter._write_standard_csv(csvfile, measurements)
            
            return True
        except Exception as e:
            print(f"CSV export failed: {e}")
            return False
    
    @staticmethod
    def _write_standard_csv(csvfile, measurements: List[MeasurementResult]):
        """Write standard CSV format"""
        fieldnames = ['timestamp', 'distance_mm', 'distance_m', 'protocol', 'type']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        for measurement in measurements:
            if measurement.type == 'measurement':
                timestamp = getattr(measurement, 'timestamp', datetime.now())
                writer.writerow({
                    'timestamp': timestamp.isoformat() if timestamp else '',
                    'distance_mm': measurement.distance_mm,
                    'distance_m': measurement.distance_m,
                    'protocol': measurement.protocol,
                    'type': measurement.type
                })
    
    @staticmethod
    def _write_door_csv(csvfile, measurements: List[MeasurementResult]):
        """Write door-specific CSV format (3 measurements per door)"""
        fieldnames = ['door_number', 'breite_mm', 'hoehe_mm', 'wandstaerke_mm', 'timestamp']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        # Group measurements by doors (every 3 measurements)
        valid_measurements = [m for m in measurements if m.type == 'measurement']
        
        door_number = 1
        for i in range(0, len(valid_measurements), 3):
            if i + 2 < len(valid_measurements):  # Ensure we have 3 measurements
                breite = valid_measurements[i]
                hoehe = valid_measurements[i + 1]
                wandstaerke = valid_measurements[i + 2]
                
                timestamp = getattr(breite, 'timestamp', datetime.now())
                writer.writerow({
                    'door_number': door_number,
                    'breite_mm': breite.distance_mm,
                    'hoehe_mm': hoehe.distance_mm,
                    'wandstaerke_mm': wandstaerke.distance_mm,
                    'timestamp': timestamp.isoformat() if timestamp else ''
                })
                door_number += 1
    
    @staticmethod
    def to_json(measurements: List[MeasurementResult], 
                filepath: str,
                door_format: bool = False) -> bool:
        """Export measurements to JSON format
        
        Args:
            measurements: List of measurement results
            filepath: Output JSON file path
            door_format: Use door-specific formatting
            
        Returns:
            True if export successful
        """
        try:
            path = Path(filepath)
            path.parent.mkdir(parents=True, exist_ok=True)
            
            if door_format:
                data = MeasurementExporter._format_door_data(measurements)
            else:
                data = MeasurementExporter._format_standard_data(measurements)
            
            with open(path, 'w', encoding='utf-8') as jsonfile:
                json.dump(data, jsonfile, indent=2, ensure_ascii=False, default=str)
            
            return True
        except Exception as e:
            print(f"JSON export failed: {e}")
            return False
    
    @staticmethod
    def _format_standard_data(measurements: List[MeasurementResult]) -> Dict[str, Any]:
        """Format data for standard JSON export"""
        export_data = {
            'export_timestamp': datetime.now().isoformat(),
            'total_measurements': len([m for m in measurements if m.type == 'measurement']),
            'measurements': []
        }
        
        for measurement in measurements:
            if measurement.type == 'measurement':
                data = asdict(measurement)
                data['timestamp'] = getattr(measurement, 'timestamp', datetime.now()).isoformat()
                # Remove raw bytes for JSON serialization
                data.pop('raw', None)
                export_data['measurements'].append(data)
        
        return export_data
    
    @staticmethod
    def _format_door_data(measurements: List[MeasurementResult]) -> Dict[str, Any]:
        """Format data for door-specific JSON export"""
        valid_measurements = [m for m in measurements if m.type == 'measurement']
        doors = []
        
        door_number = 1
        for i in range(0, len(valid_measurements), 3):
            if i + 2 < len(valid_measurements):
                breite = valid_measurements[i]
                hoehe = valid_measurements[i + 1]
                wandstaerke = valid_measurements[i + 2]
                
                door = {
                    'door_number': door_number,
                    'measurements': {
                        'breite': {
                            'distance_mm': breite.distance_mm,
                            'distance_m': breite.distance_m,
                            'timestamp': getattr(breite, 'timestamp', datetime.now()).isoformat()
                        },
                        'hoehe': {
                            'distance_mm': hoehe.distance_mm,
                            'distance_m': hoehe.distance_m,
                            'timestamp': getattr(hoehe, 'timestamp', datetime.now()).isoformat()
                        },
                        'wandstaerke': {
                            'distance_mm': wandstaerke.distance_mm,
                            'distance_m': wandstaerke.distance_m,
                            'timestamp': getattr(wandstaerke, 'timestamp', datetime.now()).isoformat()
                        }
                    }
                }
                doors.append(door)
                door_number += 1
        
        return {
            'export_timestamp': datetime.now().isoformat(),
            'total_doors': len(doors),
            'doors': doors
        }
    
    @staticmethod
    def to_excel(measurements: List[MeasurementResult], 
                 filepath: str,
                 door_format: bool = False) -> bool:
        """Export measurements to Excel format
        
        Args:
            measurements: List of measurement results
            filepath: Output Excel file path
            door_format: Use door-specific formatting
            
        Returns:
            True if export successful
        """
        if not EXCEL_AVAILABLE:
            print("Excel export requires openpyxl: pip install openpyxl")
            return False
        
        try:
            path = Path(filepath)
            path.parent.mkdir(parents=True, exist_ok=True)
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            
            if door_format:
                MeasurementExporter._write_door_excel(worksheet, measurements)
                worksheet.title = "Door Measurements"
            else:
                MeasurementExporter._write_standard_excel(worksheet, measurements)
                worksheet.title = "Measurements"
            
            workbook.save(path)
            return True
        except Exception as e:
            print(f"Excel export failed: {e}")
            return False
    
    @staticmethod
    def _write_standard_excel(worksheet, measurements: List[MeasurementResult]):
        """Write standard Excel format"""
        # Headers
        headers = ['Timestamp', 'Distance (mm)', 'Distance (m)', 'Protocol', 'Type']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Data
        row = 2
        for measurement in measurements:
            if measurement.type == 'measurement':
                timestamp = getattr(measurement, 'timestamp', datetime.now())
                worksheet.cell(row=row, column=1, value=timestamp)
                worksheet.cell(row=row, column=2, value=measurement.distance_mm)
                worksheet.cell(row=row, column=3, value=measurement.distance_m)
                worksheet.cell(row=row, column=4, value=measurement.protocol)
                worksheet.cell(row=row, column=5, value=measurement.type)
                row += 1
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _write_door_excel(worksheet, measurements: List[MeasurementResult]):
        """Write door-specific Excel format"""
        # Headers
        headers = ['Door Number', 'Breite (mm)', 'Höhe (mm)', 'Wandstärke (mm)', 'Timestamp']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Data
        valid_measurements = [m for m in measurements if m.type == 'measurement']
        row = 2
        door_number = 1
        
        for i in range(0, len(valid_measurements), 3):
            if i + 2 < len(valid_measurements):
                breite = valid_measurements[i]
                hoehe = valid_measurements[i + 1]
                wandstaerke = valid_measurements[i + 2]
                
                timestamp = getattr(breite, 'timestamp', datetime.now())
                worksheet.cell(row=row, column=1, value=door_number)
                worksheet.cell(row=row, column=2, value=breite.distance_mm)
                worksheet.cell(row=row, column=3, value=hoehe.distance_mm)
                worksheet.cell(row=row, column=4, value=wandstaerke.distance_mm)
                worksheet.cell(row=row, column=5, value=timestamp)
                row += 1
                door_number += 1
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width


def export_measurements(measurements: List[MeasurementResult],
                       filepath: str,
                       format_type: str = 'csv',
                       door_format: bool = False) -> bool:
    """Convenience function to export measurements
    
    Args:
        measurements: List of measurement results
        filepath: Output file path
        format_type: Export format ('csv', 'json', 'excel')
        door_format: Use door-specific formatting
        
    Returns:
        True if export successful
    """
    format_type = format_type.lower()
    
    if format_type == 'csv':
        return MeasurementExporter.to_csv(measurements, filepath, door_format)
    elif format_type == 'json':
        return MeasurementExporter.to_json(measurements, filepath, door_format)
    elif format_type in ('excel', 'xlsx'):
        return MeasurementExporter.to_excel(measurements, filepath, door_format)
    else:
        print(f"Unsupported format: {format_type}")
        return False


def get_default_filename(format_type: str = 'csv', door_format: bool = False) -> str:
    """Get default filename for export
    
    Args:
        format_type: Export format
        door_format: Door-specific format
        
    Returns:
        Default filename with timestamp
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    prefix = 'doors' if door_format else 'measurements'
    
    extensions = {
        'csv': 'csv',
        'json': 'json',
        'excel': 'xlsx',
        'xlsx': 'xlsx'
    }
    
    ext = extensions.get(format_type.lower(), 'csv')
    return f"{prefix}_{timestamp}.{ext}"